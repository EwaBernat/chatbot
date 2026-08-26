#!/usr/bin/env python3
"""Profiler danych dla skilla `dane-i-glos`.

Czyta CSV/TSV/XLSX/JSON/JSONL i wypisuje rzetelny profil (Markdown albo JSON),
z ktorego pisze sie scenariusz narracji. Zadna liczba w narracji nie moze
pochodzic z innego zrodla niz ten profil.

Zaleznosci: tylko biblioteka standardowa. XLSX wymaga `openpyxl`
(pip install openpyxl) — bez niego skrypt mowi to wprost.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

MAX_TOP = 5
DATE_PATTERNS = (
    "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S",
    "%d.%m.%Y %H:%M", "%Y/%m/%d",
)


# --- wczytywanie -----------------------------------------------------------

def wczytaj(sciezka: Path, arkusz: str | None) -> list[dict]:
    suf = sciezka.suffix.lower()
    if suf in (".csv", ".tsv", ".txt"):
        return _wczytaj_csv(sciezka)
    if suf in (".xlsx", ".xlsm"):
        return _wczytaj_xlsx(sciezka, arkusz)
    if suf == ".jsonl":
        return _wczytaj_jsonl(sciezka)
    if suf == ".json":
        return _wczytaj_json(sciezka)
    raise SystemExit(
        f"Nieobslugiwane rozszerzenie: {suf or '(brak)'}. "
        "Obslugiwane: .csv .tsv .xlsx .xlsm .json .jsonl"
    )


def _wczytaj_csv(sciezka: Path) -> list[dict]:
    surowe = sciezka.read_text(encoding="utf-8-sig", errors="replace")
    if not surowe.strip():
        return []
    probka = surowe[:8192]
    try:
        dialekt = csv.Sniffer().sniff(probka, delimiters=",;\t|")
    except csv.Error:
        dialekt = csv.excel
        dialekt.delimiter = "\t" if sciezka.suffix.lower() == ".tsv" else ","
    return [dict(w) for w in csv.DictReader(surowe.splitlines(), dialect=dialekt)]


def _wczytaj_xlsx(sciezka: Path, arkusz: str | None) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Do odczytu XLSX potrzebny jest openpyxl. Zainstaluj: pip install openpyxl\n"
            "Albo zapisz arkusz jako CSV i podaj ten plik."
        )
    wb = load_workbook(sciezka, read_only=True, data_only=True)
    ws = wb[arkusz] if arkusz else wb[wb.sheetnames[0]]
    wiersze = ws.iter_rows(values_only=True)
    try:
        naglowki = [str(k).strip() if k is not None else f"kolumna_{i}"
                    for i, k in enumerate(next(wiersze), 1)]
    except StopIteration:
        return []
    dane = []
    for w in wiersze:
        if all(k is None or str(k).strip() == "" for k in w):
            continue
        dane.append({n: w[i] if i < len(w) else None for i, n in enumerate(naglowki)})
    wb.close()
    return dane


def _wczytaj_jsonl(sciezka: Path) -> list[dict]:
    dane = []
    for nr, linia in enumerate(sciezka.read_text(encoding="utf-8").splitlines(), 1):
        linia = linia.strip()
        if not linia:
            continue
        try:
            obiekt = json.loads(linia)
        except json.JSONDecodeError as e:
            raise SystemExit(f"Blad JSON w linii {nr}: {e}")
        if isinstance(obiekt, dict):
            dane.append(obiekt)
    return dane


def _wczytaj_json(sciezka: Path) -> list[dict]:
    obiekt = json.loads(sciezka.read_text(encoding="utf-8"))
    if isinstance(obiekt, list):
        return [w for w in obiekt if isinstance(w, dict)]
    if isinstance(obiekt, dict):
        for wartosc in obiekt.values():
            if isinstance(wartosc, list) and wartosc and isinstance(wartosc[0], dict):
                return wartosc
        return [obiekt]
    raise SystemExit("JSON nie zawiera listy rekordow ani obiektu.")


# --- rozpoznawanie typow ---------------------------------------------------

def _pusty(w) -> bool:
    return w is None or (isinstance(w, str) and w.strip() in ("", "-", "brak", "n/d", "NA", "null"))


def _liczba(w):
    """Zwraca float albo None. Rozumie polski zapis: '1 234,56', '87,5%', '12 zl'."""
    if isinstance(w, bool):
        return None
    if isinstance(w, (int, float)):
        return None if isinstance(w, float) and math.isnan(w) else float(w)
    if not isinstance(w, str):
        return None
    t = w.strip().replace(" ", "").replace(" ", "")
    t = re.sub(r"(?i)(zl|pln|%|szt\.?|os\.?)$", "", t).strip()
    if not t:
        return None
    if "," in t and "." in t:
        t = t.replace(".", "").replace(",", ".")   # 1.234,56
    else:
        t = t.replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def _data(w):
    if isinstance(w, (datetime, date)):
        return w if isinstance(w, date) else w.date()
    if not isinstance(w, str) or not w.strip():
        return None
    t = w.strip()
    for wzor in DATE_PATTERNS:
        try:
            return datetime.strptime(t, wzor).date()
        except ValueError:
            continue
    return None


def profiluj_kolumne(nazwa: str, wartosci: list) -> dict:
    ogolem = len(wartosci)
    obecne = [w for w in wartosci if not _pusty(w)]
    braki = ogolem - len(obecne)
    wynik = {
        "kolumna": nazwa,
        "wypelnione": len(obecne),
        "braki": braki,
        "braki_proc": round(100 * braki / ogolem, 1) if ogolem else 0.0,
        "unikalne": len({str(w) for w in obecne}),
    }
    if not obecne:
        wynik["typ"] = "pusta"
        return wynik

    liczby = [_liczba(w) for w in obecne]
    if all(l is not None for l in liczby):
        wynik["typ"] = "liczbowa"
        wynik.update(
            suma=round(sum(liczby), 4),
            srednia=round(statistics.fmean(liczby), 4),
            mediana=round(statistics.median(liczby), 4),
            minimum=round(min(liczby), 4),
            maksimum=round(max(liczby), 4),
        )
        if len(liczby) > 1:
            wynik["odchylenie"] = round(statistics.stdev(liczby), 4)
        return wynik

    daty = [_data(w) for w in obecne]
    if all(d is not None for d in daty):
        wynik["typ"] = "data"
        wynik.update(od=str(min(daty)), do=str(max(daty)))
        return wynik

    wynik["typ"] = "tekstowa"
    zliczenia: dict[str, int] = {}
    for w in obecne:
        klucz = str(w).strip()
        zliczenia[klucz] = zliczenia.get(klucz, 0) + 1
    najczestsze = sorted(zliczenia.items(), key=lambda p: (-p[1], p[0]))[:MAX_TOP]
    wynik["najczestsze"] = [
        {"wartosc": k, "ile": v, "udzial_proc": round(100 * v / len(obecne), 1)}
        for k, v in najczestsze
    ]
    return wynik


def grupuj(dane: list[dict], po: str, agreguj: str | None) -> list[dict]:
    kubelki: dict[str, list] = {}
    for w in dane:
        klucz = "(brak)" if _pusty(w.get(po)) else str(w.get(po)).strip()
        kubelki.setdefault(klucz, []).append(w)
    wynik = []
    for klucz, wiersze in kubelki.items():
        pozycja = {"grupa": klucz, "liczebnosc": len(wiersze)}
        if agreguj:
            liczby = [l for l in (_liczba(w.get(agreguj)) for w in wiersze) if l is not None]
            if liczby:
                pozycja.update(
                    suma=round(sum(liczby), 4),
                    srednia=round(statistics.fmean(liczby), 4),
                    minimum=round(min(liczby), 4),
                    maksimum=round(max(liczby), 4),
                )
        wynik.append(pozycja)
    return sorted(wynik, key=lambda p: -p["liczebnosc"])


# --- prezentacja -----------------------------------------------------------

def jako_markdown(profil: dict) -> str:
    L = [f"# Profil danych: {profil['plik']}", ""]
    L.append(f"- Wierszy: **{profil['wierszy']}**")
    L.append(f"- Kolumn: **{profil['kolumn']}**")
    if profil.get("arkusz"):
        L.append(f"- Arkusz: **{profil['arkusz']}**")
    L += ["", "## Kolumny", ""]
    for k in profil["kolumny"]:
        L.append(f"### {k['kolumna']}  ·  _{k['typ']}_")
        L.append(f"- wypelnione: {k['wypelnione']}, braki: {k['braki']} ({k['braki_proc']}%), "
                 f"unikalne: {k['unikalne']}")
        if k["typ"] == "liczbowa":
            L.append(f"- suma: {k['suma']}, srednia: {k['srednia']}, mediana: {k['mediana']}")
            L.append(f"- min: {k['minimum']}, maks: {k['maksimum']}"
                     + (f", odchylenie std: {k['odchylenie']}" if "odchylenie" in k else ""))
        elif k["typ"] == "data":
            L.append(f"- zakres: {k['od']} — {k['do']}")
        elif k["typ"] == "tekstowa":
            for p in k["najczestsze"]:
                L.append(f"  - `{p['wartosc']}` — {p['ile']} ({p['udzial_proc']}%)")
        L.append("")
    if profil.get("grupy"):
        g = profil["grupy"]
        L += [f"## Przekroj wedlug `{g['po']}`"
              + (f", agregacja `{g['agreguj']}`" if g.get("agreguj") else ""), ""]
        naglowki = [k for k in g["pozycje"][0].keys()] if g["pozycje"] else []
        L.append("| " + " | ".join(naglowki) + " |")
        L.append("|" + "|".join(["---"] * len(naglowki)) + "|")
        for p in g["pozycje"]:
            L.append("| " + " | ".join(str(p.get(k, "")) for k in naglowki) + " |")
        L.append("")
    L += ["---", "Liczby do narracji bierz wylacznie z tego profilu."]
    return "\n".join(L)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Profiluje dane pod scenariusz narracji (skill dane-i-glos).")
    ap.add_argument("plik", type=Path, help="CSV / TSV / XLSX / JSON / JSONL")
    ap.add_argument("--profil", action="store_true",
                    help="(domyslne) pelny profil kolumn")
    ap.add_argument("--kolumny", help="tylko te kolumny, po przecinku")
    ap.add_argument("--grupuj", help="kolumna grupujaca")
    ap.add_argument("--agreguj", help="kolumna liczbowa agregowana w grupach")
    ap.add_argument("--arkusz", help="nazwa arkusza w XLSX (domyslnie pierwszy)")
    ap.add_argument("--json", action="store_true", help="wynik jako JSON zamiast Markdown")
    a = ap.parse_args()

    if not a.plik.exists():
        print(f"Nie ma pliku: {a.plik}", file=sys.stderr)
        return 1

    dane = wczytaj(a.plik, a.arkusz)
    if not dane:
        print(f"Plik `{a.plik}` nie zawiera zadnych wierszy danych "
              "(same naglowki albo pusty). Narracji nie ma z czego zbudowac.", file=sys.stderr)
        return 2

    kolumny = list(dane[0].keys())
    if a.kolumny:
        chciane = [k.strip() for k in a.kolumny.split(",")]
        brakujace = [k for k in chciane if k not in kolumny]
        if brakujace:
            print(f"Nie ma takich kolumn: {', '.join(brakujace)}.\n"
                  f"Dostepne: {', '.join(kolumny)}", file=sys.stderr)
            return 3
        kolumny = chciane

    profil = {
        "plik": str(a.plik),
        "wierszy": len(dane),
        "kolumn": len(dane[0]),
        "arkusz": a.arkusz,
        "kolumny": [profiluj_kolumne(k, [w.get(k) for w in dane]) for k in kolumny],
    }
    if a.grupuj:
        if a.grupuj not in dane[0]:
            print(f"Nie ma kolumny `{a.grupuj}`. Dostepne: {', '.join(dane[0].keys())}",
                  file=sys.stderr)
            return 3
        profil["grupy"] = {
            "po": a.grupuj,
            "agreguj": a.agreguj,
            "pozycje": grupuj(dane, a.grupuj, a.agreguj),
        }

    print(json.dumps(profil, ensure_ascii=False, indent=2, default=str)
          if a.json else jako_markdown(profil))
    return 0


if __name__ == "__main__":
    sys.exit(main())
